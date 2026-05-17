#!/usr/bin/env python3
"""
Run once from ~/gaza-scholarship-guide:
  python3 import_marks_data.py

Imports:
1. gaza_scholarship_dataset_100.csv → MongoDB scholarships collection
2. All knowledge files (txt, pdf, xlsx) → MongoDB knowledge collection

Data persists permanently in MongoDB Atlas — survives server restarts and Render redeploys.
"""

import csv, json, os, re, sys
from datetime import datetime

# ── Load .env ─────────────────────────────────
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if not os.path.exists(env_path):
        print("❌ .env file not found")
        sys.exit(1)
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

load_env()

MONGODB_URI = os.environ.get('MONGODB_URI')
if not MONGODB_URI:
    print("❌ MONGODB_URI not found in .env")
    sys.exit(1)

try:
    from pymongo import MongoClient, ReplaceOne
    print("✅ pymongo imported")
except ImportError:
    print("❌ pymongo not installed. Run: pip install pymongo dnspython")
    sys.exit(1)

# ── Connect to MongoDB ────────────────────────
print(f"\nConnecting to MongoDB Atlas...")
client = MongoClient(MONGODB_URI)
db = client['gazadb']
schol_col   = db['scholarships']
know_col    = db['knowledges']

# Test connection
client.admin.command('ping')
print("✅ Connected to MongoDB Atlas\n")

BASE = os.path.dirname(os.path.abspath(__file__))
KB   = os.path.join(BASE, 'knowledge_base')

# ── Country → flag emoji ──────────────────────
FLAGS = {
    'usa': '🇺🇸', 'united states': '🇺🇸', 'uk': '🇬🇧', 'united kingdom': '🇬🇧',
    'ireland': '🇮🇪', 'germany': '🇩🇪', 'france': '🇫🇷', 'turkey': '🇹🇷',
    'türkiye': '🇹🇷', 'hungary': '🇭🇺', 'malaysia': '🇲🇾', 'china': '🇨🇳',
    'japan': '🇯🇵', 'south korea': '🇰🇷', 'korea': '🇰🇷', 'russia': '🇷🇺',
    'canada': '🇨🇦', 'australia': '🇦🇺', 'new zealand': '🇳🇿', 'norway': '🇳🇴',
    'sweden': '🇸🇪', 'finland': '🇫🇮', 'denmark': '🇩🇰', 'netherlands': '🇳🇱',
    'belgium': '🇧🇪', 'switzerland': '🇨🇭', 'austria': '🇦🇹', 'italy': '🇮🇹',
    'spain': '🇪🇸', 'portugal': '🇵🇹', 'poland': '🇵🇱', 'czech': '🇨🇿',
    'romania': '🇷🇴', 'bulgaria': '🇧🇬', 'greece': '🇬🇷', 'croatia': '🇭🇷',
    'egypt': '🇪🇬', 'jordan': '🇯🇴', 'lebanon': '🇱🇧', 'morocco': '🇲🇦',
    'tunisia': '🇹🇳', 'qatar': '🇶🇦', 'kuwait': '🇰🇼', 'saudi': '🇸🇦',
    'uae': '🇦🇪', 'united arab': '🇦🇪', 'bahrain': '🇧🇭', 'oman': '🇴🇲',
    'pakistan': '🇵🇰', 'india': '🇮🇳', 'bangladesh': '🇧🇩', 'indonesia': '🇮🇩',
    'thailand': '🇹🇭', 'vietnam': '🇻🇳', 'philippines': '🇵🇭', 'singapore': '🇸🇬',
    'taiwan': '🇹🇼', 'brazil': '🇧🇷', 'south africa': '🇿🇦', 'nigeria': '🇳🇬',
    'kenya': '🇰🇪', 'multiple': '🌍', 'global': '🌍', 'various': '🌍',
    'international': '🌍', 'europe': '🇪🇺', 'european': '🇪🇺',
}

def get_flag(country):
    c = (country or '').lower().strip()
    for k, v in FLAGS.items():
        if k in c:
            return v
    return '🌍'

def slug(name):
    return re.sub(r'[^a-z0-9-]', '', re.sub(r'\s+', '-', (name or '').lower()))[:60]

def parse_funding(coverage):
    return 'partial' if 'partial' in (coverage or '').lower() else 'full'

def parse_english(notes, name):
    text = ((notes or '') + ' ' + (name or '')).lower()
    if 'ielts' in text or 'toefl' in text or 'english proficiency' in text:
        if 'waiver' in text or 'no ielts' in text or 'not required' in text:
            return False, None, True
        if 'b2' in text or '6.5' in text or '6.0' in text:
            return True, 'B2', False
        if 'b1' in text or '5.5' in text:
            return True, 'B1', False
        return True, 'B2', False
    return False, None, True

def parse_visa(country, relevance):
    c = (country or '').lower()
    if any(x in c for x in ['turkey','türkiye','jordan','malaysia','pakistan','egypt','morocco','qatar','kuwait','uae','saudi']):
        return 'high'
    if any(x in c for x in ['usa','united states','uk','united kingdom','australia','canada']):
        return 'low'
    return 'moderate'

def parse_level(degree):
    d = (degree or '').lower()
    levels = []
    if any(x in d for x in ['bachelor','undergrad','bsc','ba ']):
        levels.append('undergraduate')
    if any(x in d for x in ['master','msc','ma ','postgrad','taught','graduate']):
        levels.append('graduate')
    if any(x in d for x in ['phd','doctoral','doctorate','research']):
        levels.append('phd')
    if any(x in d for x in ['all','any','various','multiple']):
        levels = ['undergraduate','graduate','phd']
    return levels if levels else ['graduate']

def parse_fields(best_for, notes):
    text = ((best_for or '') + ' ' + (notes or '')).lower()
    if any(x in text for x in ['engineer','stem','science','technology','math','comput']):
        if any(x in text for x in ['medicine','medical','health']):
            return ['engineering','science','it','medicine']
        return ['engineering','science','it']
    if any(x in text for x in ['medicine','medical','health']):
        return ['medicine','health']
    if any(x in text for x in ['humanities','social','arts','law']):
        return ['humanities','social']
    return ['all']

def parse_gaza(relevance, target, name):
    text = ((relevance or '') + ' ' + (target or '') + ' ' + (name or '')).lower()
    return 'very high' in text or 'gaza' in text or 'palestin' in text

# ── 1. IMPORT CSV → MongoDB ───────────────────
print("=" * 50)
print("STEP 1: Importing scholarships from CSV")
print("=" * 50)

csv_path = None
for f in os.listdir(KB):
    if 'dataset_100' in f.lower() and f.endswith('.csv'):
        csv_path = os.path.join(KB, f)
        break

if not csv_path:
    print("❌ CSV file not found in knowledge_base/")
else:
    scholarships = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name     = (row.get('Opportunity_Name') or '').strip()
            country  = (row.get('Country_or_Region') or '').strip()
            degree   = (row.get('Degree_Level') or '').strip()
            coverage = (row.get('Coverage_Type') or '').strip()
            target   = (row.get('Target_Group') or '').strip()
            relevance= (row.get('Gaza_or_Palestinian_Relevance') or '').strip()
            best_for = (row.get('Best_For') or '').strip()
            url      = (row.get('Source_URL') or '').strip()
            verified = (row.get('Verification_Status') or '').strip()
            notes    = (row.get('Notes') or '').strip()
            provider = (row.get('Provider') or '').strip()

            if not name:
                continue

            eng_req, eng_min, waiver = parse_english(notes, name)

            s = {
                'id':                slug(name),
                'name':              name,
                'country':           country,
                'flag':              get_flag(country),
                'provider':          provider,
                'funding':           parse_funding(coverage),
                'covers':            coverage,
                'fields':            parse_fields(best_for, notes),
                'level':             parse_level(degree),
                'english_required':  eng_req,
                'english_min':       eng_min,
                'ielts_waiver':      waiver,
                'visa_feasibility':  parse_visa(country, relevance),
                'deadline':          'Check website',
                'gpa_min':           70,
                'link':              url,
                'required_documents':['Passport','Academic transcripts','Personal statement'],
                'notes':             notes or best_for,
                'visa_notes':        '',
                'gaza_specific':     parse_gaza(relevance, target, name),
                'verified':          'verified' in verified.lower(),
                'last_updated':      datetime.now().strftime('%Y-%m-%d'),
                'target_group':      target,
                'degree_level':      degree,
            }
            scholarships.append(s)

    # Clear existing and insert all
    schol_col.delete_many({})
    if scholarships:
        schol_col.insert_many(scholarships)
    print(f"✅ Imported {len(scholarships)} scholarships → MongoDB Atlas (scholarships collection)")

    # Also save to JSON as backup
    os.makedirs(os.path.join(BASE, 'data'), exist_ok=True)
    with open(os.path.join(BASE, 'data', 'scholarships.json'), 'w') as f:
        json.dump(scholarships, f, indent=2, ensure_ascii=False)
    print(f"✅ Backup saved → data/scholarships.json")

# ── 2. IMPORT KNOWLEDGE → MongoDB ────────────
print("\n" + "=" * 50)
print("STEP 2: Importing knowledge documents")
print("=" * 50)

knowledge_texts = []

# Text files
txt_files = [
    'Scholarship_Application_Playbook (1).txt',
    'Student_Profile_Strategies (1).txt',
    'Displacement_Documentation_Guide (1).txt',
]
for fname in txt_files:
    fpath = os.path.join(KB, fname)
    if not os.path.exists(fpath):
        print(f"⚠️  Skipping {fname} — not found")
        continue
    with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
        text = f.read().strip()
    if text:
        knowledge_texts.append({'source': fname, 'content': text})
        print(f"✅ Read: {fname} ({len(text)} chars)")

# PDF files
pdf_files = [
    'Scholarship_Essay_Templates_War_Affected_Students (1).pdf',
    'Gaza_Scholarship_Strategy_Engine (2).pdf',
    'Scholarship_Matching_Rules (1).pdf',
]
try:
    import pdfplumber
    for fname in pdf_files:
        fpath = os.path.join(KB, fname)
        if not os.path.exists(fpath):
            print(f"⚠️  Skipping {fname} — not found")
            continue
        try:
            with pdfplumber.open(fpath) as pdf:
                text = '\n'.join(p.extract_text() or '' for p in pdf.pages).strip()
            if text:
                knowledge_texts.append({'source': fname, 'content': text})
                print(f"✅ Read PDF: {fname} ({len(text)} chars)")
        except Exception as e:
            print(f"⚠️  Could not read {fname}: {e}")
except ImportError:
    print("⚠️  pdfplumber not installed — run: pip install pdfplumber")

# Excel knowledge files
xlsx_files = [
    'Palestinian_Friendly_Universities (1).xlsx',
    'Need_Met_Universities_International (1).xlsx',
    'scholarship_probability_calculator (1).xlsx',
    'gaza_scholarship_gpt_build_kit (1).xlsx',
    '100_simulated_student_profiles (1).xlsx',
]
try:
    import openpyxl
    for fname in xlsx_files:
        fpath = os.path.join(KB, fname)
        if not os.path.exists(fpath):
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            rows_text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows_text.append('\t'.join(str(c) if c is not None else '' for c in row))
            text = f"[{fname}]\n" + '\n'.join(rows_text[:300])
            if text.strip():
                knowledge_texts.append({'source': fname, 'content': text})
                print(f"✅ Read Excel: {fname}")
        except Exception as e:
            print(f"⚠️  Could not read {fname}: {e}")
except ImportError:
    print("⚠️  openpyxl not installed — run: pip install openpyxl")

# Save to MongoDB
know_col.delete_many({})
if knowledge_texts:
    know_col.insert_many(knowledge_texts)
print(f"\n✅ Saved {len(knowledge_texts)} knowledge documents → MongoDB Atlas (knowledges collection)")

# Also save JSON backup
with open(os.path.join(BASE, 'data', 'knowledge.json'), 'w') as f:
    json.dump(knowledge_texts, f, indent=2, ensure_ascii=False)
print(f"✅ Backup saved → data/knowledge.json")

# ── Summary ───────────────────────────────────
print("\n" + "=" * 50)
schol_count = schol_col.count_documents({})
know_count  = know_col.count_documents({})
print(f"🎉 Import complete!")
print(f"   Scholarships in MongoDB:   {schol_count}")
print(f"   Knowledge docs in MongoDB: {know_count}")
print(f"\n   Restart server: npm run dev")
print(f"   Check health:   curl localhost:3000/api/health")
print("=" * 50)

client.close()
